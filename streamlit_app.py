import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import os
import tempfile
import time
import zipfile
import re

EMU = 9525  

def center_image_in_cell(ws, cell, img, width, height):
    from openpyxl.utils import column_index_from_string, get_column_letter

    col_letter = ''.join(filter(str.isalpha, cell))
    row_number = int(''.join(filter(str.isdigit, cell)))

    col_index = column_index_from_string(col_letter)
    col_dim = ws.column_dimensions.get(col_letter)
    row_dim = ws.row_dimensions.get(row_number)

    col_width = col_dim.width if col_dim and col_dim.width else 8.43
    row_height = row_dim.height if row_dim and row_dim.height else 15

    cell_width_px = col_width * 7.5
    cell_height_px = row_height * 1.33

    img.width = width
    img.height = height

    img.anchor = cell

    ws.add_image(img)

def process_excel(before_file_path, template_file_path, cell_positions_dict, split_column, split_method, skiprows, sheet_name_col, barcode_folder=None, barcode_col=None, barcode_cells=None, barcode_size=(100, 50)):
    df = pd.read_excel(before_file_path, skiprows=skiprows)

    for col in df.columns:
        if df[col].dtype == 'float64':
            df[col] = df[col].fillna(0).astype('Int64')

    if split_column and split_method:
        if split_column in df.columns:
            if split_method == "Remove Numbers":
                df[split_column] = df[split_column].astype(str).str.replace(r'^\d+\s', '', regex=True)

    try:
        wb = openpyxl.load_workbook(template_file_path)
        template_sheet = wb.active
    except Exception as e:
        st.error(f"Error opening template file: {e}")
        return None

    if "Sheet1" in wb.sheetnames:
        wb.remove(wb["Sheet1"])

    template_images = []
    for image in template_sheet._images:
        temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_img_{len(template_images)}.png")
        with open(temp_img_path, "wb") as img_file:
            img_file.write(image._data())
        template_images.append((temp_img_path, image.anchor))

    for _, row in df.iterrows():
        sheet_name = str(row[sheet_name_col])[:31] if sheet_name_col in df.columns else "Default"

        if sheet_name not in wb.sheetnames:
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
        else:
            new_sheet = wb[sheet_name]

        for col_name, cell_positions in cell_positions_dict.items():
            value = row[col_name] if col_name in df.columns else ""
            for cell_pos in cell_positions:
                new_sheet[cell_pos] = value

        if barcode_folder and barcode_col and barcode_col in df.columns and barcode_cells:
            barcode_value = str(row[barcode_col])
            barcode_path_png = os.path.join(barcode_folder, f"{barcode_value}.png")
            barcode_path_jpg = os.path.join(barcode_folder, f"{barcode_value}.jpg")
            barcode_path = barcode_path_png if os.path.exists(barcode_path_png) else (barcode_path_jpg if os.path.exists(barcode_path_jpg) else None)

            if barcode_path:
                for cell in barcode_cells:
                    img = Image(barcode_path)
                    center_image_in_cell(new_sheet, cell, img, *barcode_size)

        for img_path, img_anchor in template_images:
            img = Image(img_path)
            new_sheet.add_image(img, img_anchor)

    output_path = os.path.join(tempfile.gettempdir(), "processed_excel.xlsx")
    try:
        wb.save(output_path)
    except Exception as e:
        st.error(f"Error saving the output file: {e}")
        return None

    return output_path

st.title("\U0001F4CA เว็บไซต์สร้างใบปะหน้า")
st.markdown("หมายเหตุ : หากต้องการใส่ข้อมูลหลายเซลล์ ให้ใส่เครื่องหมายจุลภาค( , )ขั้นระหว่างเซลล์")

before_file = st.file_uploader("Upload ไฟล์ข้อมูล", type=["xlsx"])
template_file = st.file_uploader("Upload ไฟล์ Template", type=["xlsx"])
barcode_zip = st.file_uploader("Upload ไฟล์ ZIP ของ Barcode (ไม่บังคับ)", type=["zip"])
skiprows = st.number_input("ระบุจำนวนแถวที่ต้องการข้าม (skiprows)", min_value=0, value=0, step=1)

if before_file:
    df = pd.read_excel(before_file, skiprows=skiprows)
    column_options = list(df.columns)
else:
    column_options = []

selected_columns = st.multiselect("เลือกคอลัมน์ที่ต้องการดึงข้อมูล", column_options)
sheet_name_col = st.selectbox("เลือกคอลัมน์สำหรับตั้งชื่อชีต", ["(ไม่เลือก)"] + column_options)

cell_positions_dict = {}
for col in selected_columns:
    cell_positions = st.text_input(f"ระบุตำแหน่งเซลล์สำหรับ {col} ", key=col)
    if cell_positions:
        cell_positions_dict[col] = [pos.strip() for pos in cell_positions.split(",")]

split_column = st.selectbox("เลือกคอลัมน์ที่ต้องการ split", ["(ไม่เลือก)"] + column_options)
split_method = st.selectbox("เลือกวิธีการ split", ["Remove Numbers", "Other Method"])

barcode_col = st.selectbox("เลือกคอลัมน์ที่ใช้จับคู่กับไฟล์ Barcode", ["(ไม่เลือก)"] + column_options)
barcode_cells_input = st.text_input("ระบุตำแหน่งเซลล์ของ Barcode (เช่น A1, B2, C3)")
barcode_cells = [re.sub(r"[^\w\d]", "", pos.strip().upper()) for pos in barcode_cells_input.split(",") if pos.strip()]
barcode_width = st.number_input("ระบุกว้างของ Barcode", min_value=10, value=100, step=10)
barcode_height = st.number_input("ระบุสูงของ Barcode", min_value=10, value=50, step=10)

barcode_folder = None
if barcode_zip:
    temp_barcode_dir = os.path.join(tempfile.gettempdir(), "barcode_images")
    os.makedirs(temp_barcode_dir, exist_ok=True)
    with zipfile.ZipFile(barcode_zip, 'r') as zip_ref:
        zip_ref.extractall(temp_barcode_dir)
    barcode_folder = temp_barcode_dir

if st.button("Generate Excel File"):
    if before_file and template_file and selected_columns and cell_positions_dict:
        temp_dir = tempfile.gettempdir()
        before_file_path = os.path.join(temp_dir, "before.xlsx")
        template_file_path = os.path.join(temp_dir, "template.xlsx")

        with open(before_file_path, "wb") as f:
            f.write(before_file.getbuffer())
        with open(template_file_path, "wb") as f:
            f.write(template_file.getbuffer())

        time.sleep(1)

        output_file = process_excel(
            before_file_path,
            template_file_path,
            cell_positions_dict,
            split_column,
            split_method,
            skiprows,
            sheet_name_col,
            barcode_folder,
            barcode_col if barcode_col != "(ไม่เลือก)" else None,
            barcode_cells,
            (barcode_width, barcode_height)
        )

        if output_file:
            with open(output_file, "rb") as file:
                st.download_button("Download Processed Excel", file, file_name="processed_excel.xlsx")
    else:
        st.error("กรุณาอัปโหลดไฟล์ เลือกคอลัมน์ และกรอกตำแหน่งเซลล์!")
